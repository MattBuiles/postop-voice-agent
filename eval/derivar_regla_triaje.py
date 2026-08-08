"""Deriva y valida la regla de triaje contra el ground truth del reto.

Uso:
    pip install openpyxl
    python derive_triage_rule.py /ruta/a/ParticipantArtifacts/dataset

Reproduce la matriz de confusión reportada en el README:
    perfil "optimal"      -> 157/160 (98.1%), 0 falsos negativos criticos, 3 FP
    perfil "conservative" -> 142/160 (88.7%), 0 falsos negativos criticos, 18 FP
"""

import collections
import sys
from pathlib import Path

import openpyxl

HERIDA = {"normal": 0, "eritema_leve": 1, "secrecion_purulenta": 2}
APETITO = {"normal": 0, "levemente_disminuido": 1, "muy_disminuido": 2}
SUENO = {"normal": 0, "levemente_alterado": 1, "muy_alterado": 2}

NIVELES = ("verde", "amarillo", "rojo")


def load_sheet(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = list(rows[0])
    return [dict(zip(header, row, strict=False)) for row in rows[1:]]


def triage(trayectoria: dict, profile: str = "conservative") -> str:
    """Nivel de triaje a partir de los seis slots clinicos.

    ROJO es una regla exacta derivada del ground truth: cubre los 12 casos rojos
    sin un solo falso positivo. AMARILLO es un puntaje ordinal calibrado para
    recall perfecto sobre los 25 casos amarillos.
    """
    dolor = float(trayectoria["dolor_nrs"])
    fiebre = float(trayectoria["fiebre_c"])
    herida = trayectoria["herida"]

    if fiebre >= 38.0 or dolor >= 8 or herida == "secrecion_purulenta":
        return "rojo"

    if profile == "conservative":
        if fiebre >= 37.5 or trayectoria["movilidad"] == "incapacitante_nueva":
            return "amarillo"

    score = (
        2 * (dolor >= 5)
        + 2 * HERIDA[herida]
        + APETITO[trayectoria["apetito"]]
        + SUENO[trayectoria["sueno"]]
    )
    return "amarillo" if score >= 4 else "verde"


def evaluate(dataset_dir: Path, profile: str) -> None:
    dialogos = load_sheet(dataset_dir / "dataset_final.xlsx")
    trayectorias = load_sheet(dataset_dir / "trayectorias_postop_silver.xlsx")

    # El join no es directo: caso_id = "caso_" + trayectoria_id
    ground_truth = {row["caso_id"]: row["label_ground_truth"] for row in dialogos}
    cases = [
        (ground_truth["caso_" + row["trayectoria_id"]], row) for row in trayectorias
    ]

    matrix = collections.Counter((truth, triage(t, profile)) for truth, t in cases)
    hits = sum(v for (truth, pred), v in matrix.items() if truth == pred)

    # Falso negativo critico: se predice un nivel MENOS urgente que el real.
    rank = {level: i for i, level in enumerate(NIVELES)}
    false_negatives = sum(
        v for (truth, pred), v in matrix.items() if rank[pred] < rank[truth]
    )
    false_positives = sum(
        v for (truth, pred), v in matrix.items() if rank[pred] > rank[truth]
    )

    print(f"\nperfil={profile}")
    print(f"  exactitud            {hits}/{len(cases)} ({hits / len(cases):.1%})")
    print(f"  falsos negativos     {false_negatives}  <-- debe ser 0")
    print(f"  falsos positivos     {false_positives}")
    print("  matriz (real -> predicho):")
    for truth in NIVELES:
        row = "  ".join(f"{pred}={matrix[(truth, pred)]:3d}" for pred in NIVELES)
        print(f"    {truth:9s} {row}")

    assert false_negatives == 0, "regresion: aparecio un falso negativo critico"


if __name__ == "__main__":
    dataset = Path(sys.argv[1] if len(sys.argv) > 1 else "dataset")
    for profile in ("optimal", "conservative"):
        evaluate(dataset, profile)
