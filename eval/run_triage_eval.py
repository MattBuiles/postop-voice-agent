"""Evalua el motor de triaje contra los 160 casos etiquetados del reto.

    .venv/bin/python eval/run_triage_eval.py

Falla con codigo distinto de cero si aparece un falso negativo critico. Eso lo
convierte en una prueba de regresion: ningun ajuste de umbrales puede colarse si
degrada la seguridad clinica.
"""

from __future__ import annotations

import argparse
import collections
import json
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ / "src"))

from postop.config import config  # noqa: E402
from postop.triage.rules import NIVELES, Slots, evaluar  # noqa: E402

_ORDEN = {nivel: i for i, nivel in enumerate(NIVELES)}


def cargar_hoja(ruta: Path) -> list[dict]:
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro[libro.sheetnames[0]]
    filas = list(hoja.iter_rows(values_only=True))
    cabecera = list(filas[0])
    return [dict(zip(cabecera, fila, strict=False)) for fila in filas[1:]]


def evaluar_perfil(casos: list[tuple[str, Slots]], perfil: str) -> dict:
    matriz: collections.Counter = collections.Counter()
    for verdad, slots in casos:
        prediccion = evaluar(slots, perfil=perfil).nivel
        matriz[(verdad, prediccion)] += 1

    aciertos = sum(v for (verdad, pred), v in matriz.items() if verdad == pred)
    # Falso negativo: se predice un nivel MENOS urgente que el real.
    falsos_negativos = sum(
        v for (verdad, pred), v in matriz.items() if _ORDEN[pred] < _ORDEN[verdad]
    )
    falsos_positivos = sum(
        v for (verdad, pred), v in matriz.items() if _ORDEN[pred] > _ORDEN[verdad]
    )
    recall_rojo = matriz[("rojo", "rojo")] / max(
        1, sum(v for (verdad, _), v in matriz.items() if verdad == "rojo")
    )
    return {
        "perfil": perfil,
        "n_casos": len(casos),
        "aciertos": aciertos,
        "exactitud": round(aciertos / len(casos), 4),
        "falsos_negativos": falsos_negativos,
        "falsos_positivos": falsos_positivos,
        "recall_rojo": round(recall_rojo, 4),
        "matriz": {f"{verdad}->{pred}": n for (verdad, pred), n in sorted(matriz.items())},
    }


def imprimir(resultado: dict) -> None:
    print(f"\nperfil = {resultado['perfil']}")
    print(f"  exactitud         {resultado['aciertos']}/{resultado['n_casos']} "
          f"({resultado['exactitud']:.1%})")
    print(f"  falsos negativos  {resultado['falsos_negativos']}   <- debe ser 0")
    print(f"  falsos positivos  {resultado['falsos_positivos']}")
    print(f"  recall de rojos   {resultado['recall_rojo']:.1%}")
    print("  matriz (real -> predicho):")
    for verdad in NIVELES:
        celdas = "  ".join(
            f"{pred}={resultado['matriz'].get(f'{verdad}->{pred}', 0):3d}" for pred in NIVELES
        )
        print(f"    {verdad:9s} {celdas}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", type=Path, help="guardar resultados en un archivo")
    args = parser.parse_args()

    dataset = config.dataset_absoluta
    dialogos = cargar_hoja(dataset / "dataset_final.xlsx")
    trayectorias = cargar_hoja(dataset / "trayectorias_postop_silver.xlsx")

    # El join no es directo: caso_id = "caso_" + trayectoria_id
    verdad_por_caso = {fila["caso_id"]: fila["label_ground_truth"] for fila in dialogos}
    casos: list[tuple[str, Slots]] = []
    for fila in trayectorias:
        verdad = verdad_por_caso["caso_" + fila["trayectoria_id"]]
        casos.append(
            (
                verdad,
                Slots(
                    dolor_nrs=float(fila["dolor_nrs"]),
                    fiebre_c=float(fila["fiebre_c"]),
                    movilidad=fila["movilidad"],
                    herida=fila["herida"],
                    apetito=fila["apetito"],
                    sueno=fila["sueno"],
                ),
            )
        )

    resultados = [evaluar_perfil(casos, perfil) for perfil in ("optimal", "conservative")]
    for resultado in resultados:
        imprimir(resultado)

    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(json.dumps(resultados, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"\nResultados en {args.json}")

    fallos = [r for r in resultados if r["falsos_negativos"] > 0]
    if fallos:
        print("\nREGRESION: hay falsos negativos criticos en "
              f"{', '.join(r['perfil'] for r in fallos)}")
        return 1
    print("\nOK: ningun falso negativo critico en ningun perfil.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
