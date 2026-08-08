"""Resumen estructurado de la llamada y persistencia de la alerta.

La rubrica pide que al terminar quede "un resumen que identifique al paciente y
su procedimiento, los sintomas reportados, la decision tomada, las referencias
usadas y los proximos pasos", y pregunta explicitamente "con que estructura y con
que persistencia". De ahi las tres salidas de este modulo:

  - un JSON de trabajo, que es lo que ve la consola;
  - un bundle **FHIR R4**, porque un resumen clinico que no habla el idioma de
    los sistemas de salud no se integra con nada;
  - filas en SQLite, para que la alerta sobreviva a la llamada y tenga estado.
"""

from __future__ import annotations

import json
import uuid
from pathlib import Path

import openpyxl

from postop.obs.traza import ahora

NOMBRE_SLOT = {
    "dolor_nrs": "Dolor (0-10)",
    "fiebre_c": "Temperatura (°C)",
    "movilidad": "Movilidad",
    "herida": "Herida quirúrgica",
    "apetito": "Apetito",
    "sueno": "Sueño",
}

PROXIMOS_PASOS = {
    "verde": "Continuar recuperación en casa. Consultar si aparecen signos de alarma.",
    "amarillo": "Contacto del equipo clínico en las próximas 24 horas.",
    "rojo": "Valoración médica inmediata. Caso escalado al equipo clínico.",
}

# LOINC para las observaciones que tienen codigo estandar. Las que no, van con
# un codigo local declarado como tal: inventar un LOINC seria peor que no ponerlo.
LOINC = {
    "dolor_nrs": ("72514-3", "Pain severity - 0-10 verbal numeric rating"),
    "fiebre_c": ("8310-5", "Body temperature"),
}


def construir(estado) -> dict:
    """Resumen de trabajo de la llamada."""
    decision = estado.decision or estado.evaluar()
    sintomas = [
        {
            "slot": slot,
            "etiqueta": NOMBRE_SLOT[slot],
            "valor": valor,
            "reportado": valor is not None,
        }
        for slot, valor in estado.slots.__dict__.items()
    ]
    citas = [cita for turno in estado.turnos for cita in turno.citas]

    return {
        "call_id": estado.call_id,
        "paciente": {
            "paciente_id": estado.paciente_id,
            "procedimiento": estado.procedimiento,
            "dia_postop": estado.dia_postop,
        },
        "sintomas": sintomas,
        "no_resueltos": estado.agotados,
        "banderas_rojas": estado.banderas,
        "participo_tercero": estado.participo_tercero,
        "decision": decision.to_dict(),
        "referencias": citas,
        "proximos_pasos": PROXIMOS_PASOS[decision.nivel],
        "n_turnos": len(estado.turnos),
        "generado_ts": ahora(),
    }


def a_fhir(resumen: dict) -> dict:
    """Bundle FHIR R4 con el encuentro, las observaciones y la alerta."""
    call_id = resumen["call_id"]
    paciente_ref = f"Patient/{resumen['paciente']['paciente_id'] or 'desconocido'}"
    entradas: list[dict] = [
        {
            "resource": {
                "resourceType": "Encounter",
                "id": call_id,
                "status": "finished",
                "class": {
                    "system": "http://terminology.hl7.org/CodeSystem/v3-ActCode",
                    "code": "VR",
                    "display": "virtual",
                },
                "subject": {"reference": paciente_ref},
                "reasonCode": [
                    {"text": f"Seguimiento postoperatorio día {resumen['paciente']['dia_postop']}"}
                ],
            }
        }
    ]

    for sintoma in resumen["sintomas"]:
        if not sintoma["reportado"]:
            continue
        slot = sintoma["slot"]
        codigo, display = LOINC.get(slot, (f"postop-{slot}", sintoma["etiqueta"]))
        sistema = "http://loinc.org" if slot in LOINC else "urn:postop-voice-agent:slots"
        recurso = {
            "resourceType": "Observation",
            "id": f"{call_id}-{slot}",
            "status": "final",
            "code": {"coding": [{"system": sistema, "code": codigo, "display": display}]},
            "subject": {"reference": paciente_ref},
            "encounter": {"reference": f"Encounter/{call_id}"},
        }
        valor = sintoma["valor"]
        if isinstance(valor, (int, float)):
            unidad = "Cel" if slot == "fiebre_c" else "{score}"
            recurso["valueQuantity"] = {"value": valor, "unit": unidad,
                                        "system": "http://unitsofmeasure.org", "code": unidad}
        else:
            recurso["valueString"] = str(valor)
        entradas.append({"resource": recurso})

    nivel = resumen["decision"]["nivel"]
    if nivel != "verde":
        entradas.append({
            "resource": {
                "resourceType": "Flag",
                "id": f"{call_id}-alerta",
                "status": "active",
                "category": [{"coding": [{
                    "system": "http://terminology.hl7.org/CodeSystem/flag-category",
                    "code": "clinical"}]}],
                "code": {"text": f"Triaje {nivel}: {resumen['proximos_pasos']}"},
                "subject": {"reference": paciente_ref},
                "encounter": {"reference": f"Encounter/{call_id}"},
            }
        })

    for i, referencia in enumerate(resumen["referencias"], 1):
        if not referencia.get("documento"):
            continue
        entradas.append({
            "resource": {
                "resourceType": "DocumentReference",
                "id": f"{call_id}-ref-{i}",
                "status": "current",
                "description": f"{referencia['documento']} p.{referencia.get('pagina')}",
                "subject": {"reference": paciente_ref},
            }
        })

    return {"resourceType": "Bundle", "type": "collection", "id": call_id, "entry": entradas}


def persistir(conn, estado) -> dict:
    """Guarda resumen, FHIR y alerta. La alerta queda con estado propio para que
    sobreviva a la llamada, que es lo que la rubrica pregunta por 'persistencia'."""
    resumen = construir(estado)
    fhir = a_fhir(resumen)

    conn.execute(
        "UPDATE calls SET fin_ts = ?, triaje_final = ?, resumen_json = ?, fhir_json = ? "
        "WHERE call_id = ?",
        (ahora(), resumen["decision"]["nivel"], json.dumps(resumen, ensure_ascii=False),
         json.dumps(fhir, ensure_ascii=False), estado.call_id),
    )

    for i, turno in enumerate(estado.turnos):
        conn.execute(
            "INSERT OR REPLACE INTO turns (turn_id, call_id, idx, hablante, texto, slots_json, "
            "citas_json, latencias_json, tokens_json, ts) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (f"{estado.call_id}-{i}", estado.call_id, i, turno.hablante, turno.texto,
             json.dumps(turno.extraccion, ensure_ascii=False),
             json.dumps(turno.citas, ensure_ascii=False),
             json.dumps(turno.latencias, ensure_ascii=False),
             json.dumps(turno.tokens, ensure_ascii=False), ahora()),
        )

    nivel = resumen["decision"]["nivel"]
    if nivel != "verde":
        existe = conn.execute(
            "SELECT alert_id FROM alerts WHERE call_id = ?", (estado.call_id,)
        ).fetchone()
        if not existe:
            conn.execute(
                "INSERT INTO alerts (alert_id, call_id, nivel, motivo_json, evidencias_json, "
                "estado, creado_ts) VALUES (?,?,?,?,?,'abierta',?)",
                (str(uuid.uuid4())[:8], estado.call_id, nivel,
                 json.dumps(resumen["decision"]["motivos"], ensure_ascii=False),
                 json.dumps([t.texto for t in estado.turnos if t.hablante != "agente"],
                            ensure_ascii=False),
                 ahora()),
            )
    conn.commit()
    return resumen


_CACHE_PACIENTES: list[dict] | None = None


def cargar_pacientes(dataset_dir: Path) -> list[dict]:
    """Perfiles del dataset del reto, para poder elegir a quien se llama."""
    global _CACHE_PACIENTES
    if _CACHE_PACIENTES is not None:
        return _CACHE_PACIENTES

    ruta = dataset_dir / "perfiles_clinicos_pacientes_silver_contest.xlsx"
    if not ruta.exists():
        _CACHE_PACIENTES = []
        return _CACHE_PACIENTES

    def hoja(archivo: Path) -> list[dict]:
        libro = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        pagina = libro[libro.sheetnames[0]]
        filas = list(pagina.iter_rows(values_only=True))
        cabecera = list(filas[0])
        return [dict(zip(cabecera, f, strict=False)) for f in filas[1:]]

    clinicos = hoja(ruta)
    demograficos = {}
    ruta_demo = dataset_dir / "perfiles_pacientes_co.xlsx"
    if ruta_demo.exists():
        demograficos = {f["paciente_id"]: f for f in hoja(ruta_demo)}

    _CACHE_PACIENTES = [
        {
            "paciente_id": c["paciente_id"],
            "nombre": demograficos.get(c["paciente_id"], {}).get("nombre_completo", c["paciente_id"]),
            "procedimiento": c["procedimiento"],
            "edad": c["edad"],
            "genero": c["genero"],
            "comorbilidades": c["comorbilidades"],
            "ciudad": demograficos.get(c["paciente_id"], {}).get("ciudad"),
        }
        for c in clinicos
    ]
    return _CACHE_PACIENTES
